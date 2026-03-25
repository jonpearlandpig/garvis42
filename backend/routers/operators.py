from fastapi import APIRouter, File, UploadFile, Form, HTTPException, Request
from fastapi.responses import StreamingResponse
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from docx import Document
from openpyxl import Workbook, load_workbook
from io import BytesIO
import uuid
from datetime import datetime, timezone
import os
from motor.motor_asyncio import AsyncIOMotorClient

# Reuse the DB connection from the environment
_mongo_client = AsyncIOMotorClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
_db = _mongo_client[os.environ.get("DB_NAME", "garvis")]

router = APIRouter(prefix="/api/operators", tags=["Operators"])

@router.post("/dochandler")
async def doc_handler(
    action: str = Form(...),  # "create", "read", "edit"
    file_type: str = Form(...),  # "pdf", "docx", "xlsx"
    file: UploadFile = File(None),  # Optional for create
    content: str = Form(None)  # For create/edit
):
    if action == "create":
        if file_type == "pdf":
            buffer = BytesIO()
            c = canvas.Canvas(buffer)
            c.drawString(100, 100, content or "Test PDF")
            c.save()
            buffer.seek(0)
            return StreamingResponse(buffer, media_type="application/pdf")
        elif file_type == "docx":
            buffer = BytesIO()
            doc = Document()
            doc.add_paragraph(content or "Test DOCX")
            doc.save(buffer)
            buffer.seek(0)
            return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        elif file_type == "xlsx":
            buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws["A1"] = content or "Test XLSX"
            wb.save(buffer)
            buffer.seek(0)
            return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # Add read/edit logic similarly using the libraries

# Subagent spawning endpoint
@router.post("/spawn")
async def spawn_subagent(task: str = Form(...), operator: str = Form(...)):
    subagent_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    record = {
        "subagent_id": subagent_id,
        "operator": operator,
        "task": task,
        "status": "queued",
        "created_at": now,
        "updated_at": now,
        "result": None,
    }
    await _db.subagent_tasks.insert_one(record)
    await _db.audit_log.insert_one({
        "log_id": str(uuid.uuid4()),
        "action": "spawn_subagent",
        "content_type": "subagent",
        "content_id": subagent_id,
        "operator": operator,
        "task": task,
        "timestamp": now,
    })
    return {
        "status": "queued",
        "subagent_id": subagent_id,
        "operator": operator,
        "task": task,
    }


@router.get("/spawn/{subagent_id}")
async def get_subagent_status(subagent_id: str):
    """Check the status of a spawned subagent task."""
    record = await _db.subagent_tasks.find_one({"subagent_id": subagent_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Subagent task not found")
    return record


@router.patch("/spawn/{subagent_id}")
async def update_subagent_status(subagent_id: str, status: str = Form(...), result: str = Form(None)):
    """Update a subagent task's status (e.g. running → complete)."""
    allowed = {"queued", "running", "complete", "failed"}
    if status not in allowed:
        raise HTTPException(status_code=400, detail=f"Status must be one of: {allowed}")
    record = await _db.subagent_tasks.find_one({"subagent_id": subagent_id})
    if not record:
        raise HTTPException(status_code=404, detail="Subagent task not found")
    now = datetime.now(timezone.utc).isoformat()
    await _db.subagent_tasks.update_one(
        {"subagent_id": subagent_id},
        {"$set": {"status": status, "result": result, "updated_at": now}}
    )
    return {"subagent_id": subagent_id, "status": status}
