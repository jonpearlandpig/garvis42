"""
TourText Emergent — FastAPI Router
SMS-first tour information system integrated into the Garvis platform.

Collections (MongoDB):
  tt_tours          — active tours
  tt_source_files   — uploaded source documents
  tt_truth_records  — verified canonical data
  tt_invocations    — append-only query audit log
  tt_escalations    — unresolved / low-confidence tickets
"""
import asyncio
import hashlib
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, Request, UploadFile
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field

from ..tourtext_integrations import openai_processor, supabase_storage, twilio_client
from ..tourtext_utils import (
    generate_inv_taid,
    generate_rec_taid,
    generate_src_taid,
    generate_tkt_taid,
    generate_tour_code,
    generate_tour_taid,
    generate_uuid,
    hash_file,
    hash_phone_number,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/tourtext", tags=["TourText"])

# ---------------------------------------------------------------------------
# MongoDB — reuse the same connection string as the main app
# ---------------------------------------------------------------------------
_mongo_url = os.environ.get("MONGO_URL", os.environ.get("MONGO_URI", "mongodb://mongo:27017"))
_db_name = os.environ.get("DB_NAME", "garvis")

_tt_client = AsyncIOMotorClient(_mongo_url)
tt_db = _tt_client[_db_name]


# ---------------------------------------------------------------------------
# Pydantic request / response models
# ---------------------------------------------------------------------------

class TourCreate(BaseModel):
    name: str
    start_date: Optional[str] = None
    multi_tour_access: bool = False


class TourOut(BaseModel):
    tour_id: str
    taid: str
    name: str
    tour_code: str
    start_date: Optional[str]
    status: str
    created_at: str


class QueryRequest(BaseModel):
    tour_code: str
    query: str
    phone: Optional[str] = None


class QueryResponse(BaseModel):
    answer: str
    answer_policy: str
    confidence: float
    invocation_taid: str
    escalation_taid: Optional[str] = None


class EscalationCreate(BaseModel):
    tour_id: str
    description: str
    severity: str = "medium"
    assigned_role: str = "PM"


class TruthRecordCreate(BaseModel):
    tour_id: str
    record_type: str  # show | venue | finance | safety | people | vip
    content: Dict[str, Any]
    source_taid: Optional[str] = None
    confidence: float = 1.0


# ---------------------------------------------------------------------------
# Helper: keyword scorer
# ---------------------------------------------------------------------------

def _score_record(record: Dict, keywords: List[str]) -> float:
    content_str = str(record.get("content", "")).lower()
    record_type = record.get("record_type", "").lower()
    hits = sum(1 for kw in keywords if kw.lower() in content_str or kw.lower() in record_type)
    return hits / max(len(keywords), 1)


# ---------------------------------------------------------------------------
# Background file processor
# ---------------------------------------------------------------------------

async def _process_uploaded_file(file_id: str, filename: str, content: bytes, file_type: str, tour_id: str):
    """Extract truth records from an uploaded file and persist them."""
    try:
        truth_records = []

        if filename.lower().endswith(".csv") or file_type == "text/csv":
            import csv, io
            reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
            for row in reader:
                truth_records.append({
                    "record_type": "show",
                    "content": dict(row),
                    "confidence": 0.85,
                })

        elif filename.lower().endswith((".xlsx", ".xls")):
            try:
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    record = dict(zip(headers, [str(v) if v is not None else "" for v in row]))
                    truth_records.append({"record_type": "show", "content": record, "confidence": 0.85})
            except Exception as exc:
                logger.warning("Excel parse error: %s", exc)

        elif filename.lower().endswith(".pdf") or file_type == "application/pdf":
            try:
                import pdfplumber, io, re
                STOP_RE = re.compile(
                    r"STOP\s*#?(\d+)\s+(\w+ \d{1,2})\s+([^,]+),\s*([A-Z]{2})\s+(.+)", re.IGNORECASE
                )
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        for line in text.splitlines():
                            m = STOP_RE.match(line.strip())
                            if m:
                                truth_records.append({
                                    "record_type": "show",
                                    "content": {
                                        "stop": m.group(1),
                                        "date": m.group(2),
                                        "city": m.group(3).strip(),
                                        "state": m.group(4),
                                        "venue": m.group(5).strip(),
                                    },
                                    "confidence": 0.9,
                                })
            except Exception as exc:
                logger.warning("PDF parse error: %s", exc)

        now = datetime.now(timezone.utc).isoformat()
        for tr in truth_records:
            doc = {
                "taid": generate_rec_taid(),
                "record_id": generate_uuid(),
                "tour_id": tour_id,
                "source_file_id": file_id,
                "record_type": tr["record_type"],
                "content": tr["content"],
                "confidence": tr["confidence"],
                "status": "verified",
                "financial_guardrail": tr["record_type"] == "finance",
                "created_at": now,
                "updated_at": now,
            }
            await tt_db.tt_truth_records.insert_one(doc)

        await tt_db.tt_source_files.update_one(
            {"file_id": file_id},
            {"$set": {"processing_status": "completed", "record_count": len(truth_records), "updated_at": now}},
        )
        logger.info("Processed %s → %d truth records", filename, len(truth_records))

    except Exception as exc:
        logger.error("File processing failed for %s: %s", file_id, exc)
        await tt_db.tt_source_files.update_one(
            {"file_id": file_id},
            {"$set": {"processing_status": "failed", "error": str(exc)}},
        )


# ---------------------------------------------------------------------------
# Tour management
# ---------------------------------------------------------------------------

@router.post("/tours", response_model=TourOut, status_code=201)
async def create_tour(payload: TourCreate):
    tour_id = generate_uuid()
    taid = generate_tour_taid()
    tour_code = generate_tour_code(payload.name)
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "tour_id": tour_id,
        "taid": taid,
        "name": payload.name,
        "tour_code": tour_code,
        "start_date": payload.start_date,
        "multi_tour_access": payload.multi_tour_access,
        "status": "active",
        "created_at": now,
        "updated_at": now,
    }
    await tt_db.tt_tours.insert_one(doc)
    return TourOut(**{k: v for k, v in doc.items() if k != "_id"})


@router.get("/tours", response_model=List[TourOut])
async def list_tours():
    cursor = tt_db.tt_tours.find({"status": "active"}, {"_id": 0})
    return [TourOut(**t) async for t in cursor]


@router.get("/tours/{tour_id}", response_model=TourOut)
async def get_tour(tour_id: str):
    tour = await tt_db.tt_tours.find_one({"tour_id": tour_id}, {"_id": 0})
    if not tour:
        raise HTTPException(status_code=404, detail="Tour not found")
    return TourOut(**tour)


# ---------------------------------------------------------------------------
# File upload
# ---------------------------------------------------------------------------

@router.post("/tours/{tour_id}/upload")
async def upload_file(
    tour_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    file_type: str = Form("other"),
):
    tour = await tt_db.tt_tours.find_one({"tour_id": tour_id})
    if not tour:
        raise HTTPException(status_code=404, detail="Tour not found")

    content = await file.read()
    file_hash = hash_file(content)

    existing = await tt_db.tt_source_files.find_one({"tour_id": tour_id, "file_hash": file_hash})
    if existing:
        return {"file_id": existing["file_id"], "status": "duplicate", "message": "File already uploaded"}

    file_id = generate_uuid()
    taid = generate_src_taid()
    storage_filename = f"{tour_id}/{file_id}_{file.filename}"
    now = datetime.now(timezone.utc).isoformat()

    upload_result = await supabase_storage.upload(storage_filename, content)

    doc = {
        "file_id": file_id,
        "taid": taid,
        "tour_id": tour_id,
        "filename": file.filename,
        "file_type": file_type,
        "content_type": file.content_type or "application/octet-stream",
        "size": len(content),
        "file_hash": file_hash,
        "storage_path": upload_result.get("path", ""),
        "storage_provider": upload_result.get("provider", "local"),
        "processing_status": "pending",
        "record_count": 0,
        "created_at": now,
        "updated_at": now,
    }
    await tt_db.tt_source_files.insert_one(doc)

    background_tasks.add_task(
        _process_uploaded_file, file_id, file.filename, content, file.content_type or "", tour_id
    )

    return {"file_id": file_id, "taid": taid, "status": "processing", "filename": file.filename}


@router.get("/tours/{tour_id}/files")
async def list_files(tour_id: str):
    cursor = tt_db.tt_source_files.find({"tour_id": tour_id}, {"_id": 0})
    return [f async for f in cursor]


@router.post("/files/{file_id}/reprocess")
async def reprocess_file(file_id: str, background_tasks: BackgroundTasks):
    src = await tt_db.tt_source_files.find_one({"file_id": file_id})
    if not src:
        raise HTTPException(status_code=404, detail="Source file not found")

    storage_path = src.get("storage_path", "")
    # Attempt to read back from local storage if available
    try:
        content = Path(storage_path).read_bytes()
    except Exception:
        raise HTTPException(status_code=422, detail="File content not available for reprocessing")

    await tt_db.tt_source_files.update_one({"file_id": file_id}, {"$set": {"processing_status": "pending"}})
    background_tasks.add_task(
        _process_uploaded_file, file_id, src["filename"], content, src.get("content_type", ""), src["tour_id"]
    )
    return {"file_id": file_id, "status": "reprocessing"}


# ---------------------------------------------------------------------------
# Truth records
# ---------------------------------------------------------------------------

@router.get("/tours/{tour_id}/truth-records")
async def list_truth_records(tour_id: str, record_type: Optional[str] = None):
    query: Dict[str, Any] = {"tour_id": tour_id}
    if record_type:
        query["record_type"] = record_type
    cursor = tt_db.tt_truth_records.find(query, {"_id": 0})
    return [r async for r in cursor]


@router.post("/truth-records", status_code=201)
async def create_truth_record(payload: TruthRecordCreate):
    tour = await tt_db.tt_tours.find_one({"tour_id": payload.tour_id})
    if not tour:
        raise HTTPException(status_code=404, detail="Tour not found")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "taid": generate_rec_taid(),
        "record_id": generate_uuid(),
        "tour_id": payload.tour_id,
        "record_type": payload.record_type,
        "content": payload.content,
        "confidence": payload.confidence,
        "source_file_id": payload.source_taid,
        "status": "verified",
        "financial_guardrail": payload.record_type == "finance",
        "created_at": now,
        "updated_at": now,
    }
    await tt_db.tt_truth_records.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


# ---------------------------------------------------------------------------
# Query processing pipeline
# ---------------------------------------------------------------------------

@router.post("/query", response_model=QueryResponse)
async def process_query(payload: QueryRequest):
    tour = await tt_db.tt_tours.find_one({"tour_code": payload.tour_code, "status": "active"})
    if not tour:
        raise HTTPException(status_code=404, detail=f"No active tour with code '{payload.tour_code}'")

    tour_id = tour["tour_id"]
    invocation_taid = generate_inv_taid()
    now = datetime.now(timezone.utc).isoformat()
    phone_hash = hash_phone_number(payload.phone) if payload.phone else None

    # 1. Parse intent
    intent_data = await openai_processor.parse_intent(payload.query)
    keywords: List[str] = intent_data.get("keywords", payload.query.lower().split()[:8])
    intent: str = intent_data.get("intent", "general")

    # 2. Retrieve truth records for this tour
    all_records = await tt_db.tt_truth_records.find(
        {"tour_id": tour_id, "status": "verified"}, {"_id": 0}
    ).to_list(length=200)

    # 3. Score and rank
    scored = [(r, _score_record(r, keywords)) for r in all_records]
    scored.sort(key=lambda x: x[1], reverse=True)
    top_records = [r for r, score in scored if score > 0][:5]
    confidence = scored[0][1] if scored else 0.0

    # 4. Financial guardrail
    if intent == "finance":
        answer = (
            "Financial queries require explicit approval. "
            "Please contact your Tour Manager for settlement details."
        )
        answer_policy = "refusal"
        escalation_taid = await _create_escalation(
            tour_id, "Finance query requires guardrail approval", payload.query, "high", "Finance"
        )
        await _log_invocation(
            invocation_taid, tour_id, phone_hash, payload.query, intent,
            answer_policy, confidence, answer, escalation_taid
        )
        return QueryResponse(
            answer=answer,
            answer_policy=answer_policy,
            confidence=0.0,
            invocation_taid=invocation_taid,
            escalation_taid=escalation_taid,
        )

    # 5. Route by confidence
    escalation_taid = None
    if confidence >= 0.5 and top_records:
        answer = await openai_processor.format_response(payload.query, top_records)
        answer_policy = "truth_record"
    elif top_records:
        answer = await openai_processor.format_response(payload.query, top_records)
        answer_policy = "normalized"
        escalation_taid = await _create_escalation(
            tour_id, f"Low confidence ({confidence:.2f}) for query", payload.query, "low", "PM"
        )
    else:
        answer = (
            "I don't have verified information for that query. "
            "An escalation ticket has been created for your Tour Manager."
        )
        answer_policy = "escalate"
        escalation_taid = await _create_escalation(
            tour_id, "No matching truth records found", payload.query, "medium", "PM"
        )

    await _log_invocation(
        invocation_taid, tour_id, phone_hash, payload.query, intent,
        answer_policy, confidence, answer, escalation_taid
    )

    return QueryResponse(
        answer=answer,
        answer_policy=answer_policy,
        confidence=confidence,
        invocation_taid=invocation_taid,
        escalation_taid=escalation_taid,
    )


async def _create_escalation(
    tour_id: str, description: str, original_query: str, severity: str, assigned_role: str
) -> str:
    taid = generate_tkt_taid()
    now = datetime.now(timezone.utc).isoformat()
    await tt_db.tt_escalations.insert_one({
        "taid": taid,
        "escalation_id": generate_uuid(),
        "tour_id": tour_id,
        "description": description,
        "original_query": original_query,
        "severity": severity,
        "assigned_role": assigned_role,
        "status": "open",
        "created_at": now,
        "updated_at": now,
    })
    return taid


async def _log_invocation(
    taid: str, tour_id: str, phone_hash: Optional[str], query: str,
    intent: str, answer_policy: str, confidence: float, response: str,
    escalation_taid: Optional[str],
):
    now = datetime.now(timezone.utc).isoformat()
    await tt_db.tt_invocations.insert_one({
        "taid": taid,
        "invocation_id": generate_uuid(),
        "tour_id": tour_id,
        "phone_hash": phone_hash,
        "query_text": query,
        "query_intent": intent,
        "answer_policy": answer_policy,
        "confidence": confidence,
        "response_text": response,
        "escalation_taid": escalation_taid,
        "created_at": now,
    })


# ---------------------------------------------------------------------------
# SMS webhook (Twilio)
# ---------------------------------------------------------------------------

@router.post("/sms/webhook")
async def sms_webhook(request: Request):
    """Twilio SMS webhook — receives inbound texts and replies via TourText query pipeline."""
    form = await request.form()
    from_number: str = form.get("From", "")
    body: str = form.get("Body", "").strip()
    tour_code: str = form.get("ToCity", "")  # Can be pre-configured per Twilio number

    if not body:
        return {"status": "empty"}

    # Allow tour code as first word of message: "ROCKTR26 where is load-in?"
    parts = body.split(None, 1)
    if len(parts) == 2 and len(parts[0]) <= 8 and parts[0].isupper():
        tour_code = parts[0]
        body = parts[1]

    if not tour_code:
        await twilio_client.send_sms(
            from_number,
            "TourText: Please start your message with your tour code (e.g. ROCK26 where is load-in?)",
        )
        return {"status": "no_tour_code"}

    try:
        result = await process_query(QueryRequest(tour_code=tour_code, query=body, phone=from_number))
        await twilio_client.send_sms(from_number, result.answer)
        return {"status": "sent", "invocation_taid": result.invocation_taid}
    except HTTPException as exc:
        await twilio_client.send_sms(from_number, f"TourText: {exc.detail}")
        return {"status": "error", "detail": exc.detail}


# ---------------------------------------------------------------------------
# Escalation management
# ---------------------------------------------------------------------------

@router.get("/tours/{tour_id}/escalations")
async def list_escalations(tour_id: str, status: Optional[str] = None):
    query: Dict[str, Any] = {"tour_id": tour_id}
    if status:
        query["status"] = status
    cursor = tt_db.tt_escalations.find(query, {"_id": 0})
    return [e async for e in cursor]


@router.patch("/escalations/{taid}/resolve")
async def resolve_escalation(taid: str):
    now = datetime.now(timezone.utc).isoformat()
    result = await tt_db.tt_escalations.update_one(
        {"taid": taid},
        {"$set": {"status": "resolved", "resolved_at": now, "updated_at": now}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Escalation not found")
    return {"taid": taid, "status": "resolved"}


# ---------------------------------------------------------------------------
# Integration status / health
# ---------------------------------------------------------------------------

@router.get("/integrations/status")
async def integration_status():
    twilio_ok = bool(
        os.environ.get("TWILIO_ACCOUNT_SID")
        and os.environ.get("TWILIO_AUTH_TOKEN")
        and os.environ.get("TWILIO_PHONE_NUMBER")
    )
    openai_ok = bool(os.environ.get("OPENAI_API_KEY"))
    supabase_ok = bool(os.environ.get("SUPABASE_URL") and os.environ.get("SUPABASE_KEY"))

    tour_count = await tt_db.tt_tours.count_documents({"status": "active"})
    record_count = await tt_db.tt_truth_records.count_documents({})
    invocation_count = await tt_db.tt_invocations.count_documents({})

    return {
        "tourtext_version": "4.1",
        "integrations": {
            "twilio_sms": "configured" if twilio_ok else "not_configured",
            "openai_nlp": "configured" if openai_ok else "not_configured (basic mode)",
            "supabase_storage": "configured" if supabase_ok else "not_configured (local fallback)",
        },
        "stats": {
            "active_tours": tour_count,
            "truth_records": record_count,
            "total_invocations": invocation_count,
        },
    }
